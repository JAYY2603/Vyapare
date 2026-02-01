from datetime import datetime
from pathlib import Path
from uuid import uuid4

from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404
from django.utils.text import get_valid_filename
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from drf_spectacular.utils import extend_schema

from .models import PredictionDataset, Prediction
from .serializers import GeneratePredictionSerializer, PredictionResponseSerializer

EXPECTED_COLUMNS = [
    "Date",
    "Day",
    "OrderID",
    "Item",
    "Quantity",
    "Price",
    "Category",
    "PaymentType",
]

MAX_FILE_SIZE_BYTES = 10 * 1024 * 1024
ALLOWED_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}


def _is_non_empty_text(value):
    return isinstance(value, str) and bool(value.strip())


def _is_valid_date_cell(value):
    if isinstance(value, datetime):
        return True
    if not isinstance(value, str):
        return False

    try:
        datetime.strptime(value.strip(), "%d/%m/%Y/%H:%M:%S")
        return True
    except ValueError:
        return False


def _is_valid_quantity(value):
    if isinstance(value, bool):
        return False

    if isinstance(value, int):
        return value > 0

    if isinstance(value, float):
        return value.is_integer() and value > 0

    return False


def _is_valid_price(value):
    if isinstance(value, bool):
        return False
    return isinstance(value, (int, float)) and value >= 0


def _validate_file_type(uploaded_file):
    extension = Path(uploaded_file.name).suffix.lower()
    if extension in ALLOWED_EXTENSIONS:
        return True, "File type is valid Excel format."

    return False, "Only Excel files (.xlsx/.xlsm/.xltx/.xltm) are allowed."


def _validate_file_size(uploaded_file):
    if uploaded_file.size < MAX_FILE_SIZE_BYTES:
        return True, "File size is less than 10MB."

    return False, "File size must be less than 10MB."


def _validate_schema_and_datatypes(uploaded_file):
    try:
        workbook = load_workbook(uploaded_file, data_only=True, read_only=True)
    except Exception:
        return False, "Unable to read Excel file. The file might be corrupted."

    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)

    try:
        headers = next(rows)
    except StopIteration:
        workbook.close()
        return False, "Excel file is empty."

    normalized_headers = [
        str(cell).strip() if cell is not None else "" for cell in headers]
    if normalized_headers[: len(EXPECTED_COLUMNS)] != EXPECTED_COLUMNS:
        workbook.close()
        return (
            False,
            "Required columns are missing or in incorrect order. Expected: "
            + ", ".join(EXPECTED_COLUMNS),
        )

    has_data_row = False
    for row_number, row in enumerate(rows, start=2):
        relevant_cells = list(row[: len(EXPECTED_COLUMNS)])

        if not any(cell is not None and str(cell).strip() != "" for cell in relevant_cells):
            continue

        has_data_row = True

        validators = [
            ("Date", _is_valid_date_cell),
            ("Day", _is_non_empty_text),
            ("OrderID", _is_non_empty_text),
            ("Item", _is_non_empty_text),
            ("Quantity", _is_valid_quantity),
            ("Price", _is_valid_price),
            ("Category", _is_non_empty_text),
            ("PaymentType", _is_non_empty_text),
        ]

        for index, (column_name, validator) in enumerate(validators):
            value = relevant_cells[index] if index < len(
                relevant_cells) else None
            if not validator(value):
                workbook.close()
                return (
                    False,
                    f"Invalid datatype/value for '{column_name}' at row {row_number}.",
                )

    workbook.close()

    if not has_data_row:
        return False, "Excel file must include at least one data row."

    return True, "Required columns and datatypes are valid."


def _save_uploaded_file(uploaded_file, user_id):
    upload_root = Path(settings.DATASET_STORAGE_ROOT) / \
        "prediction_datasets" / str(user_id)
    upload_root.mkdir(parents=True, exist_ok=True)

    safe_name = get_valid_filename(uploaded_file.name)
    saved_name = f"{uuid4().hex[:8]}_{safe_name}"
    target_path = upload_root / saved_name

    with target_path.open("wb+") as destination:
        for chunk in uploaded_file.chunks():
            destination.write(chunk)

    return str(target_path)


@login_required
def upload_dataset(request):
    validation_results = []
    upload_success = False
    uploaded_file_name = None
    dataset_id = None

    if request.method == "POST":
        uploaded_file = request.FILES.get("dataset_file")

        if uploaded_file is None:
            validation_results = [
                {
                    "title": "File selected",
                    "is_valid": False,
                    "message": "Please choose an Excel file to upload.",
                }
            ]
        else:
            uploaded_file_name = uploaded_file.name

            is_type_valid, type_message = _validate_file_type(uploaded_file)
            validation_results.append(
                {
                    "title": "1. File Type: Excel",
                    "is_valid": is_type_valid,
                    "message": type_message,
                }
            )

            is_size_valid, size_message = _validate_file_size(uploaded_file)
            validation_results.append(
                {
                    "title": "2. File Size: < 10MB",
                    "is_valid": is_size_valid,
                    "message": size_message,
                }
            )

            if is_type_valid and is_size_valid:
                is_schema_valid, schema_message = _validate_schema_and_datatypes(
                    uploaded_file)
            else:
                is_schema_valid = False
                schema_message = "Skipped because file type or file size validation failed."

            validation_results.append(
                {
                    "title": "3. Required Columns + Datatypes",
                    "is_valid": is_schema_valid,
                    "message": schema_message,
                }
            )

            if is_type_valid and is_size_valid and is_schema_valid:
                file_path = _save_uploaded_file(uploaded_file, request.user.id)
                prediction_dataset = PredictionDataset.objects.create(
                    user=request.user,
                    file_name=uploaded_file.name,
                    file_path=file_path,
                    is_validated=True,
                )
                dataset_id = prediction_dataset.id
                upload_success = True

    context = {
        "validation_results": validation_results,
        "upload_success": upload_success,
        "uploaded_file_name": uploaded_file_name,
        "dataset_id": dataset_id,
        "required_columns": EXPECTED_COLUMNS,
    }
    return render(request, "predictions/upload_dataset.html", context)


@extend_schema(
    request=GeneratePredictionSerializer,
    responses={200: PredictionResponseSerializer},
    tags=["Predictions"],
    description="Generate predictions for uploaded dataset.",
)
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def api_generate_predictions(request):
    """
    Generate predictions for a validated dataset.
    For now, returns a boilerplate prediction.
    """
    serializer = GeneratePredictionSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(
            {
                "message": "Please fix the highlighted errors.",
                "errors": serializer.errors,
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    dataset_id = serializer.validated_data['dataset_id']
    dataset = get_object_or_404(
        PredictionDataset,
        id=dataset_id,
        user=request.user,
        is_validated=True
    )

    # Generate boilerplate prediction result
    boilerplate_result = {
        "summary": "Prediction Analysis Complete",
        "status": "success",
        "metrics": {
            "accuracy": 0.92,
            "confidence": 0.87,
            "data_points": 1000,
        },
        "insights": [
            "The dataset shows a positive trend across all categories.",
            "Peak performance observed during weekends.",
            "Electronic category has the highest sales volume.",
        ],
        "recommendations": [
            "Increase inventory for high-demand items.",
            "Optimize pricing during peak hours.",
            "Focus marketing on top-performing categories.",
        ],
    }

    prediction = Prediction.objects.create(
        dataset=dataset,
        prediction_result=boilerplate_result,
    )

    return Response(
        {
            "message": "Predictions generated successfully.",
            "prediction_id": prediction.id,
            "prediction_result": prediction.prediction_result,
            "generated_at": prediction.generated_at,
        },
        status=status.HTTP_200_OK,
    )
