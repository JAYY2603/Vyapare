from datetime import datetime
from pathlib import Path
from uuid import uuid4

from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.utils.text import get_valid_filename
from openpyxl import load_workbook

EXPECTED_COLUMNS = [
    "Date",
    "Day",
    "OrderID",
    "Item",
    "Quantity",
    "Price",
    "Category",
    "PaymentType",
]

MAX_FILE_SIZE_BYTES = 10 * 1024 * 1024
ALLOWED_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}


def _is_non_empty_text(value):
    return isinstance(value, str) and bool(value.strip())


def _is_valid_date_cell(value):
    if isinstance(value, datetime):
        return True
    if not isinstance(value, str):
        return False

    try:
        datetime.strptime(value.strip(), "%d/%m/%Y/%H:%M:%S")
        return True
    except ValueError:
        return False


def _is_valid_quantity(value):
    if isinstance(value, bool):
        return False

    if isinstance(value, int):
        return value > 0

    if isinstance(value, float):
        return value.is_integer() and value > 0

    return False


def _is_valid_price(value):
    if isinstance(value, bool):
        return False
    return isinstance(value, (int, float)) and value >= 0


# Validation 1: ensure uploaded file is an Excel file extension.
def _validate_file_type(uploaded_file):
    extension = Path(uploaded_file.name).suffix.lower()
    if extension in ALLOWED_EXTENSIONS:
        return True, "File type is valid Excel format."

    return False, "Only Excel files (.xlsx/.xlsm/.xltx/.xltm) are allowed."


# Validation 2: ensure uploaded file size is less than 10MB.
def _validate_file_size(uploaded_file):
    if uploaded_file.size < MAX_FILE_SIZE_BYTES:
        return True, "File size is less than 10MB."

    return False, "File size must be less than 10MB."


# Validation 3: ensure required columns exist and each column value matches expected datatype.
def _validate_schema_and_datatypes(uploaded_file):
    try:
        workbook = load_workbook(uploaded_file, data_only=True, read_only=True)
    except Exception:
        return False, "Unable to read Excel file. The file might be corrupted."

    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)

    try:
        headers = next(rows)
    except StopIteration:
        workbook.close()
        return False, "Excel file is empty."

    normalized_headers = [
        str(cell).strip() if cell is not None else "" for cell in headers]
    if normalized_headers[: len(EXPECTED_COLUMNS)] != EXPECTED_COLUMNS:
        workbook.close()
        return (
            False,
            "Required columns are missing or in incorrect order. Expected: "
            + ", ".join(EXPECTED_COLUMNS),
        )

    has_data_row = False
    for row_number, row in enumerate(rows, start=2):
        relevant_cells = list(row[: len(EXPECTED_COLUMNS)])

        if not any(cell is not None and str(cell).strip() != "" for cell in relevant_cells):
            continue

        has_data_row = True

        validators = [
            ("Date", _is_valid_date_cell),
            ("Day", _is_non_empty_text),
            ("OrderID", _is_non_empty_text),
            ("Item", _is_non_empty_text),
            ("Quantity", _is_valid_quantity),
            ("Price", _is_valid_price),
            ("Category", _is_non_empty_text),
            ("PaymentType", _is_non_empty_text),
        ]

        for index, (column_name, validator) in enumerate(validators):
            value = relevant_cells[index] if index < len(
                relevant_cells) else None
            if not validator(value):
                workbook.close()
                return (
                    False,
                    f"Invalid datatype/value for '{column_name}' at row {row_number}.",
                )

    workbook.close()

    if not has_data_row:
        return False, "Excel file must include at least one data row."

    return True, "Required columns and datatypes are valid."


def _save_uploaded_file(uploaded_file, user_id):
    upload_root = Path(settings.DATASET_STORAGE_ROOT) / \
        "uploaded_datasets" / str(user_id)
    upload_root.mkdir(parents=True, exist_ok=True)

    safe_name = get_valid_filename(uploaded_file.name)
    saved_name = f"{uuid4().hex[:8]}_{safe_name}"
    target_path = upload_root / saved_name

    with target_path.open("wb+") as destination:
        for chunk in uploaded_file.chunks():
            destination.write(chunk)

    return str(target_path), saved_name


@login_required
def upload_dataset(request):
    validation_results = []
    upload_success = False
    uploaded_file_name = None
    generated_dataset_key = None

    if request.method == "POST":
        uploaded_file = request.FILES.get("dataset_file")

        if uploaded_file is None:
            validation_results = [
                {
                    "title": "File selected",
                    "is_valid": False,
                    "message": "Please choose an Excel file to upload.",
                }
            ]
        else:
            uploaded_file_name = uploaded_file.name

            is_type_valid, type_message = _validate_file_type(uploaded_file)
            validation_results.append(
                {
                    "title": "1. File Type: Excel",
                    "is_valid": is_type_valid,
                    "message": type_message,
                }
            )

            is_size_valid, size_message = _validate_file_size(uploaded_file)
            validation_results.append(
                {
                    "title": "2. File Size: < 10MB",
                    "is_valid": is_size_valid,
                    "message": size_message,
                }
            )

            if is_type_valid and is_size_valid:
                is_schema_valid, schema_message = _validate_schema_and_datatypes(
                    uploaded_file)
            else:
                is_schema_valid = False
                schema_message = "Skipped because file type or file size validation failed."

            validation_results.append(
                {
                    "title": "3. Required Columns + Datatypes",
                    "is_valid": is_schema_valid,
                    "message": schema_message,
                }
            )

            if is_type_valid and is_size_valid and is_schema_valid:
                _, generated_dataset_key = _save_uploaded_file(
                    uploaded_file, request.user.id)
                upload_success = True

    context = {
        "validation_results": validation_results,
        "upload_success": upload_success,
        "uploaded_file_name": uploaded_file_name,
        "generated_dataset_key": generated_dataset_key,
        "required_columns": EXPECTED_COLUMNS,
    }
    return render(request, "analytics/upload_dataset.html", context)


@login_required
def generated_analytics(request, dataset_key=None, dataset_name=None):
    resolved_dataset_key = dataset_key or dataset_name or ""
    dataset_label = request.GET.get(
        "label", "").strip() or resolved_dataset_key
    context = {
        "dataset_name": dataset_label,
        "dataset_key": resolved_dataset_key,
    }
    return render(
        request,
        "analytics/generated_analytics/generated_analytics.html",
        context,
    )
