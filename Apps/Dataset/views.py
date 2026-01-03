from pathlib import Path
from datetime import datetime
from uuid import uuid4

from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.http import FileResponse, Http404
from django.shortcuts import get_object_or_404, render
from django.utils import timezone
from drf_spectacular.utils import extend_schema
from openpyxl import Workbook, load_workbook
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from .models import Dataset, InventoryItem
from .serializers import (
    CreateInventoryItemResponseSerializer,
    CreateInventoryItemSerializer,
    CreateDatasetResponseSerializer,
    CreateDatasetSerializer,
    InitInventoryResponseSerializer,
    InventoryListResponseSerializer,
    InitSaleResponseSerializer,
    RecordSaleResponseSerializer,
    RecordSaleSerializer,
    UpdateInventoryItemResponseSerializer,
    UpdateInventoryItemSerializer,
)


DATASET_COLUMNS = [
    "Date",
    "Day",
    "OrderID",
    "Item",
    "Quantity",
    "Price",
    "Category",
    "PaymentType",
]


def _create_dataset_excel_file(dataset_name):
    storage_root = Path(settings.DATASET_STORAGE_ROOT)
    storage_root.mkdir(parents=True, exist_ok=True)

    safe_name = "".join(ch if ch.isalnum() or ch in (
        "-", "_") else "_" for ch in dataset_name)
    file_name = f"{safe_name}_{uuid4().hex[:8]}.xlsx"
    file_path = storage_root / file_name

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sales"
    worksheet.append(DATASET_COLUMNS)
    workbook.save(file_path)

    return file_name, str(file_path)


@login_required
def dataset_list(request):
    datasets = Dataset.objects.filter(user=request.user)
    return render(request, "dataset_list.html", {"datasets": datasets})


@extend_schema(
    request=CreateDatasetSerializer,
    responses={201: CreateDatasetResponseSerializer},
    tags=["Dataset"],
    description="Create a new dataset and initialize an Excel file with required columns.",
)
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def api_create_dataset(request):
    serializer = CreateDatasetSerializer(
        data=request.data, context={"request": request})
    if not serializer.is_valid():
        return Response(
            {
                "message": "Please fix the highlighted errors.",
                "errors": serializer.errors,
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    dataset_name = serializer.validated_data["dataset_name"]
    file_name, file_path = _create_dataset_excel_file(dataset_name)

    dataset = Dataset.objects.create(
        user=request.user,
        name=dataset_name,
        file_name=file_name,
        file_path=file_path,
    )

    return Response(
        {
            "message": "Dataset Created",
            "dataset_id": dataset.id,
            "dataset_name": dataset.name,
            "created_at": dataset.created_at,
        },
        status=status.HTTP_201_CREATED,
    )


@login_required
def export_dataset_excel(request, dataset_id):
    dataset = get_object_or_404(Dataset, id=dataset_id, user=request.user)
    file_path = Path(dataset.file_path)

    if not file_path.exists():
        raise Http404("Dataset file not found.")

    return FileResponse(
        open(file_path, "rb"),
        as_attachment=True,
        filename=dataset.file_name,
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )


@extend_schema(
    responses={200: InitSaleResponseSerializer},
    tags=["Dataset"],
    description="Initialize a new sale session with current timestamp and next order ID.",
)
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def api_init_sale(request, dataset_id):
    dataset = get_object_or_404(Dataset, id=dataset_id, user=request.user)
    current_datetime = datetime.now().strftime("%d/%m/%Y/%H:%M:%S")
    next_order_id = f"O{dataset.last_order_number + 1}"

    return Response(
        {
            "order_id": next_order_id,
            "current_datetime": current_datetime,
        },
        status=status.HTTP_200_OK,
    )


@extend_schema(
    request=RecordSaleSerializer,
    responses={201: RecordSaleResponseSerializer},
    tags=["Dataset"],
    description="Record a sale with one or more items and append rows to the selected dataset Excel file.",
)
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def api_record_sale(request, dataset_id):
    serializer = RecordSaleSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(
            {
                "message": "Please fix the highlighted errors.",
                "errors": serializer.errors,
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    validated = serializer.validated_data
    file_path = None

    with transaction.atomic():
        dataset = Dataset.objects.select_for_update().filter(
            id=dataset_id,
            user=request.user,
        ).first()

        if dataset is None:
            return Response(
                {
                    "message": "Dataset not found.",
                },
                status=status.HTTP_404_NOT_FOUND,
            )

        dataset.last_order_number += 1
        dataset.save(update_fields=["last_order_number"])
        order_id = f"O{dataset.last_order_number}"
        file_path = Path(dataset.file_path)

    if not file_path.exists():
        return Response(
            {
                "message": "Dataset file not found.",
            },
            status=status.HTTP_404_NOT_FOUND,
        )

    workbook = load_workbook(file_path)
    worksheet = workbook.active

    recorded_at = datetime.now()
    date_text = recorded_at.strftime("%d/%m/%Y/%H:%M:%S")
    day_text = recorded_at.strftime("%A")
    payment_type = validated["payment_type"]

    for item in validated["items"]:
        worksheet.append(
            [
                date_text,
                day_text,
                order_id,
                item["item"],
                item["quantity"],
                float(item["price"]),
                item["category"],
                payment_type,
            ]
        )

    workbook.save(file_path)

    return Response(
        {
            "message": "Sale record saved successfully.",
            "order_id": order_id,
            "item_count": len(validated["items"]),
            "recorded_at": date_text,
        },
        status=status.HTTP_201_CREATED,
    )


@extend_schema(
    responses={200: InitInventoryResponseSerializer},
    tags=["Dataset"],
    description="Initialize an inventory item form with next ItemID for the selected dataset.",
)
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def api_init_inventory(request, dataset_id):
    dataset = get_object_or_404(Dataset, id=dataset_id, user=request.user)
    next_item_id = f"I{dataset.last_inventory_item_number + 1}"

    return Response(
        {
            "item_id": next_item_id,
        },
        status=status.HTTP_200_OK,
    )


@extend_schema(
    request=CreateInventoryItemSerializer,
    responses={201: CreateInventoryItemResponseSerializer},
    tags=["Dataset"],
    description="Create an inventory item for the selected dataset.",
)
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def api_create_inventory(request, dataset_id):
    dataset = get_object_or_404(Dataset, id=dataset_id, user=request.user)
    serializer = CreateInventoryItemSerializer(
        data=request.data,
        context={"request": request, "dataset": dataset},
    )

    if not serializer.is_valid():
        return Response(
            {
                "message": "Please fix the highlighted errors.",
                "errors": serializer.errors,
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    validated = serializer.validated_data

    with transaction.atomic():
        locked_dataset = Dataset.objects.select_for_update().filter(
            id=dataset.id,
            user=request.user,
        ).first()

        if locked_dataset is None:
            return Response(
                {
                    "message": "Dataset not found.",
                },
                status=status.HTTP_404_NOT_FOUND,
            )

        expected_item_id = f"I{locked_dataset.last_inventory_item_number + 1}"
        item_id = validated["item_id"]

        if item_id != expected_item_id:
            return Response(
                {
                    "message": "ItemID mismatch. Please reopen Add Inventory popup and try again.",
                },
                status=status.HTTP_409_CONFLICT,
            )

        InventoryItem.objects.create(
            dataset=locked_dataset,
            item_id=item_id,
            item_name=validated["item_name"],
            item_category=validated["item_category"],
            barcode_number=validated.get("barcode_number", ""),
            cost_price=validated["cost_price"],
            selling_price=validated["selling_price"],
        )

        locked_dataset.last_inventory_item_number += 1
        locked_dataset.save(update_fields=["last_inventory_item_number"])

    return Response(
        {
            "message": "Inventory item saved successfully.",
            "item_id": validated["item_id"],
            "item_name": validated["item_name"],
        },
        status=status.HTTP_201_CREATED,
    )


@extend_schema(
    responses={200: InventoryListResponseSerializer},
    tags=["Dataset"],
    description="Fetch inventory items for a selected dataset.",
)
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def api_list_inventory(request, dataset_id):
    dataset = get_object_or_404(Dataset, id=dataset_id, user=request.user)
    items = dataset.inventory_items.all().order_by("-created_at")

    payload = [
        {
            "inventory_id": item.id,
            "item_id": item.item_id,
            "item_name": item.item_name,
            "item_category": item.item_category,
            "barcode_number": item.barcode_number,
            "cost_price": item.cost_price,
            "selling_price": item.selling_price,
            "created_at": timezone.localtime(item.created_at).strftime("%d/%m/%Y %H:%M"),
        }
        for item in items
    ]

    return Response(
        {
            "items": payload,
        },
        status=status.HTTP_200_OK,
    )


@extend_schema(
    request=UpdateInventoryItemSerializer,
    responses={200: UpdateInventoryItemResponseSerializer},
    tags=["Dataset"],
    description="Update an inventory item for the selected dataset.",
)
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def api_update_inventory(request, dataset_id, inventory_id):
    dataset = get_object_or_404(Dataset, id=dataset_id, user=request.user)
    inventory_item = get_object_or_404(
        InventoryItem,
        id=inventory_id,
        dataset=dataset,
    )

    serializer = UpdateInventoryItemSerializer(
        data=request.data,
        context={
            "dataset": dataset,
            "inventory_item": inventory_item,
        },
    )

    if not serializer.is_valid():
        return Response(
            {
                "message": "Please fix the highlighted errors.",
                "errors": serializer.errors,
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    validated = serializer.validated_data
    inventory_item.item_name = validated["item_name"]
    inventory_item.item_category = validated["item_category"]
    inventory_item.barcode_number = validated.get("barcode_number", "")
    inventory_item.cost_price = validated["cost_price"]
    inventory_item.selling_price = validated["selling_price"]
    inventory_item.save(
        update_fields=[
            "item_name",
            "item_category",
            "barcode_number",
            "cost_price",
            "selling_price",
        ]
    )

    return Response(
        {
            "message": "Inventory item updated successfully.",
            "inventory_id": inventory_item.id,
        },
        status=status.HTTP_200_OK,
    )
